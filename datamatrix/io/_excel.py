# -*- coding: utf-8 -*-

"""
This file is part of datamatrix.

datamatrix is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

datamatrix is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with datamatrix.  If not, see <http://www.gnu.org/licenses/>.
"""


import os
from datamatrix.py3compat import *
from datamatrix import DataMatrix, MixedColumn
from datamatrix._datamatrix._seriescolumn import _SeriesColumn


def readxlsx(path, default_col_type=MixedColumn):

	"""
	desc: |
		Reads a DataMatrix from an Excel 2010 xlsx file.
		
		__Example:__
		
		~~~.python 
		dm = io.readxlsx('data.xlsx')
		~~~		

	arguments:
		path:	The path to the xlsx file.

	keywords:
		default_col_type:	The default column type.

	returns:
		A DataMatrix.
	"""

	from openpyxl import load_workbook

	wb = load_workbook(path)
	rows = list(wb.active.rows)
	dm = DataMatrix(default_col_type=default_col_type, length=len(rows)-1)
	column_names = []
	for cell in rows.pop(0):
		if cell.value is None:
			raise ValueError(u'Not all columns have a name on the first row')
		dm[cell.value] = default_col_type
		column_names.append(cell.value)
	for i, row in enumerate(rows):
		for colname, cell in zip(column_names, rows[i]):
			if cell.value is None:
				dm[colname][i] = default_col_type.default_value
				warn(u'Some rows miss column %s' % colname)
			else:
				dm[colname][i] = cell.value
	return dm


def writexlsx(dm, path):

	"""
	desc: |
		Writes a DataMatrix to an Excel 2010 xlsx file. The first sheet will
		contain a regular table with all non-series columns. SeriesColumns are
		saved as individual sheets.
		
		__Example:__
		
		~~~ .python				
		io.writexlsx(dm, 'data.xlsx')
		~~~		

	arguments:
		dm:		The DataMatrix to write.
		path:	The path to the xlsx file.
	"""

	from openpyxl import Workbook, utils

	try:
		os.makedirs(os.path.dirname(path))
	except:
		pass
	wb = Workbook()
	# First we will write all flat columns to the first sheet
	ws = wb.active
	ws.title = u'Main sheet'
	flat_columns = [colname for colname, column in dm.columns \
		if not isinstance(column, _SeriesColumn)]
	for colnr, colname in enumerate(flat_columns):
		ws[utils.get_column_letter(colnr+1)+'1'] = colname
	for rownr, row in enumerate(dm):
		for colnr, colname in enumerate(flat_columns):
			value = row[colname]
			ws[utils.get_column_letter(colnr+1)+str(rownr+2)] = safe_str(value)
	# Next we will write all series to individual sheets
	series_columns = [colname for colname, column in dm.columns \
		if isinstance(column, _SeriesColumn)]
	for colname in series_columns:
		ws = wb.create_sheet(title=colname)
		for rownr, row in enumerate(dm[colname]):
			for colnr, value in enumerate(row):
				ws[utils.get_column_letter(colnr+1)+str(rownr+1)] = \
					safe_str(value)
	wb.save(path)
